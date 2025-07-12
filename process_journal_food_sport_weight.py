import pandas as pd
import json
import re
from openpyxl import load_workbook
from datetime import datetime, timedelta


def extract_sheets_from_excel(file_path):
    """
    Extracts the 'Journal' and 'Nutritional Values' sheets from an Excel file,
    preserving formulas as strings in the 'Journal' sheet.

    Args:
        file_path (str): The path to the Excel file.

    Returns:
        tuple: A tuple containing two DataFrames: (journal_df, nutrition_df).
               Returns (None, None) if the sheets are not found.
    """
    try:
        wb = load_workbook(filename=file_path, data_only=False)
        xls = pd.ExcelFile(file_path, engine="openpyxl")
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return None, None

    journal_df = None
    nutrition_df = None

    if "Journal" in wb.sheetnames:
        ws = wb["Journal"]
        header = [cell.value for cell in ws[1]]
        cols_to_use = ["Date", "Nourriture", "Pds", "Sport"]

        if not all(col in header for col in cols_to_use):
            print(
                f"Error reading 'Journal' sheet. Check if columns {cols_to_use} exist."
            )
            journal_df = None
        else:
            data = []
            for row in ws.iter_rows(min_row=2):
                row_data = {}
                for col_name in cols_to_use:
                    col_idx = header.index(col_name)
                    cell_value = row[col_idx].value
                    row_data[col_name] = cell_value
                data.append(row_data)

            journal_df = pd.DataFrame(data, columns=cols_to_use)
            journal_df.rename(
                columns={
                    "Nourriture": "recorded food",
                    "Pds": "weight",
                    "Sport": "sport",
                },
                inplace=True,
            )
    else:
        print("Sheet 'Journal' not found.")

    if "Variables" in xls.sheet_names:
        try:
            nutrition_df = pd.read_excel(
                xls,
                sheet_name="Variables",
                engine="openpyxl",
            )
        except Exception as e:
            print(f"Error reading 'Variables' sheet: {e}")
    else:
        print("Sheet 'Variables' not found.")

    return journal_df, nutrition_df


def process_date_column(df: pd.DataFrame) -> pd.DataFrame:
    """
    Processes the 'Date' column in a DataFrame, handling initial date strings
    and subsequent formula strings like '=A2+1'.
    """
    if "Date" not in df.columns:
        raise ValueError("DataFrame must contain a 'Date' column.")

    processed_dates = []
    previous_date = None

    for index, row in df.iterrows():
        date_value = row["Date"]

        if index == 0:
            # First row contains the initial date string
            try:
                # pd.to_datetime is more robust for the first date
                previous_date = pd.to_datetime(date_value)
            except (ValueError, TypeError):
                raise ValueError(f"Invalid date format in the first row: {date_value}.")
            processed_dates.append(previous_date)
        else:
            # Subsequent rows contain formulas like '=A2+1'
            if str(date_value).strip().startswith("="):
                if previous_date is None:
                    raise ValueError("Previous date not set for formula processing.")
                previous_date += timedelta(days=1)
                processed_dates.append(previous_date)
            else:
                # If it's not a formula, try to parse it as a date
                try:
                    current_date = pd.to_datetime(date_value)
                    previous_date = current_date
                    processed_dates.append(current_date)
                except (ValueError, TypeError):
                    raise ValueError(
                        f"Unexpected value in 'Date' column at row {index}: {date_value}. Expected formula or date."
                    )

    df["Date"] = processed_dates
    return df


def transform_sport_formula(formula):
    """
    Transforms a sport formula string by replacing weight references and
    weight lifting patterns.
    """
    if not isinstance(formula, str):
        return formula

    # Replace weight reference (e.g., F316) with WEIGHT
    formula = re.sub(r"F\d+", "WEIGHT", formula)

    # Replace weight lifting pattern (e.g., 14*8) with weight_lifting(14)
    formula = re.sub(r"(\d+)\*8", r"weight_lifting(\1)", formula)

    return formula


if __name__ == "__main__":
    file_path = "Journal nutrition.xlsx"
    journal_df, nutrition_df = extract_sheets_from_excel(file_path)

    if journal_df is not None:
        journal_df = process_date_column(journal_df)

        # Filter rows where Date is after 2024-06-30
        journal_df = journal_df[journal_df["Date"] > "2024-06-30"]

        # Apply the transformation to the 'sport' column
        journal_df["sport"] = journal_df["sport"].apply(transform_sport_formula)

        print("--- Processed Journal Data ---")
        print(journal_df.head())

        journal_df.to_json(
            "journal.json", orient="records", indent=2, default_handler=str
        )
